#-*- coding: utf-8 -*-
# @Time    : 2020/3/13 10:35
# @Author  : tenghaiqin
# @File    : do_excel.py

from openpyxl import load_workbook #引入openpyxl类
from common.publice.log import Log #引入日志类
from common.publice import project_path #引入路径模块
from common.publice.read_config import ReadConfig #引入读取配置文件类

#我们根据读取配置文件的caseid 来控制读取哪一条用例

log = Log() #创建日志实例
class  Do_Excel():

    def __init__(self,file_name,sheet_name):
        '''读取Excel文件数据的类
        file_name:文件名及路径，
        sheet_name:表单名'''
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self,section):
        '''读取数据'''
        #打开工作薄
        try:
            wb = load_workbook(self.file_name,)
            sheet = wb[self.sheet_name] #打开哪张表  sheet1 sheet2
            CaseId = ReadConfig(project_path.conf_path).get_str(section,'case_id')
        except Exception as e :
            log.error('文件打开错误{}'.format(e))
            raise e #抛出异常

        all_data=[] #存储所有的数据
        #遍历行
        for i in range(2,sheet.max_row+1):
            row_data = {}
            row_data['CaseId']=sheet.cell(i,1).value
            row_data['Module']=sheet.cell(i,2).value
            row_data['Title']=sheet.cell(i,3).value
            row_data['Method']=sheet.cell(i,4).value
            row_data['Url']=sheet.cell(i,5).value
            row_data['Headers']=sheet.cell(i,6).value
            row_data['Param']=sheet.cell(i,7).value
            row_data['ExpectedResult']=sheet.cell(i,8).value
            all_data.append(row_data)

        wb.close()
        finally_data = [] #最终列表 空列表
        if CaseId == 'all': #配置文件获取值为all，就获取sheet表所有测试数据
            finally_data = all_data
        else:
            try:
                for i in eval(CaseId):#否则就遍历caseid列表 [1,2,3]获取指定 的测试数据
                    finally_data.append(all_data[i-1]) # 如：[{'用例1:'值1'},'用例2:'值2']  当用例i=2  ,all_data[i-1]==2-1=1(索引值),='用例2:'值2'
            except Exception as e:
                log.error('配置文件的CaseId值错误:{}'.format(e))

        return finally_data


    def write_back(self,row,col,value):
        '''测试结果写回excel
        row:写入的行
        col：写入的列'''
        try:
            wb = load_workbook(self.file_name)
            sheet = wb[self.sheet_name] #打开哪张表  sheet1 sheet2
        except Exception as e :
            log.error('文件打开错误{}'.format(e))
            raise e #抛出异常
        try:
            sheet.cell(row=row,column=col).value = value
            wb.save(self.file_name)#保存
            wb.close()#关闭
        except Exception as e:
            log.error('写入出错{}'.format(e))
            raise e #抛出异常



if __name__ == '__main__':
    #加个r 转义
    data = Do_Excel(project_path.case_path,'carInfo').read_data('CaseId')
    print(data)
    #先来读取一下carinfo的用例,






